# -*- coding: utf-8 -*-
"""Generación de reportes en Excel (openpyxl)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

COLOR_ENCABEZADO = "1F3A5F"
FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF")
RELLENO_ENCABEZADO = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")


def _escribir_encabezados(hoja, encabezados, fila=1):
    for columna, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila, column=columna, value=texto)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO
        celda.alignment = Alignment(horizontal="center")


def _ajustar_ancho_columnas(hoja):
    for columna in hoja.columns:
        longitud_maxima = max((len(str(celda.value)) for celda in columna if celda.value is not None), default=10)
        letra_columna = columna[0].column_letter
        hoja.column_dimensions[letra_columna].width = min(max(longitud_maxima + 2, 12), 40)


def exportar_libro_remuneraciones(liquidaciones, ruta_archivo):
    """Exporta el libro de remuneraciones (lista de liquidaciones) a un archivo .xlsx."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Libro Remuneraciones"

    encabezados = [
        "Código Empleado", "Año", "Mes", "Sueldo Base", "Ingreso Bruto",
        "Total Previsional", "Impuesto Único", "Sueldo Líquido", "Costo Empresa",
    ]
    _escribir_encabezados(hoja, encabezados)

    for fila, liquidacion in enumerate(liquidaciones, start=2):
        hoja.cell(row=fila, column=1, value=liquidacion.get("codigo_empleado"))
        hoja.cell(row=fila, column=2, value=liquidacion.get("anio"))
        hoja.cell(row=fila, column=3, value=liquidacion.get("mes"))
        hoja.cell(row=fila, column=4, value=liquidacion.get("sueldo_base"))
        hoja.cell(row=fila, column=5, value=liquidacion.get("total_ingreso_bruto"))
        hoja.cell(row=fila, column=6, value=liquidacion.get("total_descuentos_previsionales"))
        hoja.cell(row=fila, column=7, value=liquidacion.get("impuesto_unico"))
        hoja.cell(row=fila, column=8, value=liquidacion.get("sueldo_liquido"))
        hoja.cell(row=fila, column=9, value=liquidacion.get("costo_total_empresa"))

    _ajustar_ancho_columnas(hoja)
    libro.save(ruta_archivo)
    return ruta_archivo


def exportar_liquidacion_individual(liquidacion, trabajador, ruta_archivo):
    """Exporta la liquidación individual de un trabajador a un archivo .xlsx."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Liquidación"

    hoja["A1"] = "LIQUIDACIÓN DE SUELDO"
    hoja["A1"].font = Font(bold=True, size=14)
    hoja["A2"] = "Trabajador: {} {}".format(trabajador.get("nombres", ""), trabajador.get("apellido_paterno", ""))
    hoja["A3"] = "RUT: {}".format(trabajador.get("rut", ""))
    hoja["A4"] = "Período: {}/{}".format(liquidacion.get("mes"), liquidacion.get("anio"))

    filas = [
        ("Sueldo Base", liquidacion.get("sueldo_base")),
        ("Total Haberes Imponibles", liquidacion.get("total_haberes_imponibles")),
        ("Total Haberes No Imponibles", liquidacion.get("total_haberes_no_imponibles")),
        ("Total Ingreso Bruto", liquidacion.get("total_ingreso_bruto")),
        ("Descuento AFP", liquidacion.get("monto_afp")),
        ("Descuento Salud", liquidacion.get("monto_salud")),
        ("Descuento Seguro Cesantía", liquidacion.get("monto_seguro_cesantia")),
        ("Total Descuentos Previsionales", liquidacion.get("total_descuentos_previsionales")),
        ("Base Tributable", liquidacion.get("base_tributable")),
        ("Impuesto Único", liquidacion.get("impuesto_unico")),
        ("Total Descuentos", liquidacion.get("total_descuentos")),
        ("SUELDO LÍQUIDO", liquidacion.get("sueldo_liquido")),
        ("Aportes Patronales", liquidacion.get("total_aportes_patronales")),
        ("Costo Total Empresa", liquidacion.get("costo_total_empresa")),
    ]

    fila_inicio = 6
    _escribir_encabezados(hoja, ["Concepto", "Monto"], fila=fila_inicio)
    for indice, (concepto, monto) in enumerate(filas, start=fila_inicio + 1):
        hoja.cell(row=indice, column=1, value=concepto)
        hoja.cell(row=indice, column=2, value=monto)

    _ajustar_ancho_columnas(hoja)
    libro.save(ruta_archivo)
    return ruta_archivo


def exportar_ficha_trabajador(trabajador, ruta_archivo):
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Ficha Trabajador"
    _escribir_encabezados(hoja, ["Campo", "Valor"])
    for fila, (campo, valor) in enumerate(trabajador.items(), start=2):
        hoja.cell(row=fila, column=1, value=campo)
        hoja.cell(row=fila, column=2, value=str(valor) if valor is not None else "")
    _ajustar_ancho_columnas(hoja)
    libro.save(ruta_archivo)
    return ruta_archivo


def exportar_tabla_generica(titulo, encabezados, filas, ruta_archivo):
    """Exporta una tabla genérica (lista de listas/tuplas) a Excel; usada por varios informes."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo[:31] if titulo else "Reporte"
    _escribir_encabezados(hoja, encabezados)
    for fila_indice, fila in enumerate(filas, start=2):
        for columna_indice, valor in enumerate(fila, start=1):
            hoja.cell(row=fila_indice, column=columna_indice, value=valor)
    _ajustar_ancho_columnas(hoja)
    libro.save(ruta_archivo)
    return ruta_archivo
